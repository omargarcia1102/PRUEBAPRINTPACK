import MySQLdb.cursors
from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for
from app import mysql
import os
import requests
from datetime import timedelta
import MySQLdb
from werkzeug.security import generate_password_hash, check_password_hash
import random
from datetime import datetime
import io
import openpyxl
from flask import send_file


def enviar_telegram(mensaje):
    token = os.environ.get('TELEGRAM_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')
    if not token or not chat_id:
        return
    try:
        requests.post(
            f'https://api.telegram.org/bot{token}/sendMessage',
            json={
                'chat_id': chat_id,
                'text': mensaje,
                'parse_mode': 'HTML'
            },
            timeout=5
        )
    except Exception:
        pass


def hora_colombia():
    from datetime import datetime
    return (datetime.utcnow() + timedelta(hours=-5)).strftime('%d/%m/%Y, %I:%M:%S %p')


def nivel_stock(stock):
    try:
        s = int(stock)
        if s == 0:
            return "🔴 SIN STOCK"
        if s <= 50:
            return "🟡 STOCK BAJO"
        return "🟢 STOCK ALTO"
    except:
        return "⚪ N/A"

main = Blueprint('main', __name__)

# PAGINAS 


@main.route('/')
def inicio():
    return render_template('INICIO.html')


@main.route('/acceder')
def acceder():
    return render_template('ACCEDER.html')


@main.route('/inventario')
def inventario():
    if session.get('rol') != 'admin':
        return redirect(url_for('main.acceder'))
    return render_template('INVENTARIO.html')

# Admin y asesor
@main.route('/ventas')
def ventas():
    if session.get('rol') not in ['admin', 'asesor']:
        return redirect(url_for('main.acceder'))
    return render_template('VENTAS.html')


@main.route('/productos')
def productos():
    return render_template('PRODUCTOS.html')


@main.route('/procesos')
def procesos():
    return render_template('PROCESOS.html')


@main.route('/contacto')
def contacto():
    return render_template('CONTACTO.html')


# LOGIN 
from werkzeug.security import check_password_hash

@main.route('/login', methods=['POST'])
def login():
    data     = request.get_json()
    usuario  = data.get('usuario')
    password = data.get('password')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(
        "SELECT * FROM usuarios WHERE usuario = %s AND estado = 'activo'",
        (usuario,)
    )
    user = cursor.fetchone()
    cursor.close()

    if user and check_password_hash(user['password'], password):
        session['usuario']  = user['usuario']
        session['rol']      = user['rol']
        session['es_admin'] = user['rol'] == 'admin'
        return jsonify({'ok': True, 'rol': user['rol']})
    else:
        return jsonify({'ok': False, 'mensaje': 'Credenciales incorrectas'}), 401

# LOGOUT
@main.route('/logout')
def logout():
    # 1. Destruye absolutamente todas las variables de la sesión (usuario, rol, es_admin)
    session.clear()
    
    # 2. Redirige al inicio
    response = redirect(url_for('main.inicio'))
    
    # 3. Elimina la cookie del navegador por seguridad extra
    response.delete_cookie('session', path='/')
    return response

# INVENTARIO CRUD 

@main.route('/api/productos', methods=['GET'])
def get_productos():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) 
    cursor.execute("SELECT * FROM productos")
    resultado = cursor.fetchall()
    cursor.close()
    return jsonify(resultado)


@main.route('/api/productos', methods=['POST'])
def crear_producto():
    d = request.get_json()
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        INSERT INTO productos 
        (nombre, codigo, tipo, categoria, capas, espesor, material,
         color, dimensiones, peso, stock, unidad, bodega, proveedor,
         costo, fecha_ingreso, notas)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        d.get('nombre'), d.get('codigo'), d.get('tipo'),
        d.get('categoria'), d.get('capas'), d.get('espesor'),
        d.get('material'), d.get('color'), d.get('dimensiones'),
        d.get('peso'), d.get('stock'), d.get('unidad'),
        d.get('bodega'), d.get('proveedor'), d.get('costo'),
        d.get('fecha'), d.get('notas')
    ))
    
    id_nuevo = cursor.lastrowid  

   
    cursor.execute("""
        INSERT INTO movimientos 
        (id_producto, nombre_producto, tipo_movimiento, usuario, detalle, stock_anterior, stock_nuevo)
        VALUES (%s, %s, 'AGREGAR', %s, %s, %s, %s)
    """, (
        id_nuevo,
        d.get('nombre'),
        session.get('usuario'),
        f"Producto agregado al inventario",
        0,
        d.get('stock')
    ))
    alerta = nivel_stock(d.get('stock'))
    precio = f"${int(float(d.get('costo') or 0)):,}".replace(',', '.')
    mensaje = (
        f"🟢 <b>PRODUCTO AGREGADO</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 <b>Producto:</b> {d.get('nombre')}\n"
        f"👤 <b>Usuario:</b> {session.get('usuario')}\n"
        f"🏷️ <b>Tipo movimiento:</b> AGREGAR\n"
        f"📊 <b>Stock nuevo:</b> {d.get('stock')} unidades\n"
        f"💰 <b>Precio:</b> {precio} COP\n"
        f"📍 <b>Bodega:</b> {d.get('bodega') or 'No especificada'}\n"
        f"⏰ <b>Hora:</b> {hora_colombia()}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"{alerta}"
        )
    enviar_telegram(mensaje)

    mysql.connection.commit()
    cursor.close()
    return jsonify({'ok': True}), 201

@main.route('/api/productos/<int:id>', methods=['PUT'])
def actualizar_producto(id):
    d = request.get_json()
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("SELECT nombre, stock FROM productos WHERE id=%s", (id,))
    producto_actual = cursor.fetchone()
    stock_anterior = producto_actual['stock'] if producto_actual else 0
    nombre_producto = producto_actual['nombre'] if producto_actual else ''

    
    cursor.execute("""
        UPDATE productos SET
        nombre=%s, codigo=%s, tipo=%s, capas=%s,
        espesor=%s, material=%s, color=%s, stock=%s, bodega=%s,
        costo=%s, fecha_ingreso=%s, notas=%s
        WHERE id=%s
    """, (
        d.get('nombre'), d.get('codigo'), d.get('tipo'),
        d.get('capas'), d.get('espesor'), d.get('material'), d.get('color'),
        d.get('stock'), d.get('bodega'), d.get('costo'),
        d.get('fecha'), d.get('notas'), id
    ))

    cursor.execute("""
        INSERT INTO movimientos 
        (id_producto, nombre_producto, tipo_movimiento, usuario, detalle, stock_anterior, stock_nuevo)
        VALUES (%s, %s, 'EDITAR', %s, %s, %s, %s)
    """, (
        id,
        nombre_producto,
        session.get('usuario'),
        f"Producto editado",
        stock_anterior,
        d.get('stock')
    ))
    
    alerta = nivel_stock(d.get('stock'))
    precio = f"${int(float(d.get('costo') or 0)):,}".replace(',', '.')
    mensaje = (
        f"🟡 <b>PRODUCTO EDITADO</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 <b>Producto:</b> {nombre_producto}\n"
        f"👤 <b>Usuario:</b> {session.get('usuario')}\n"
        f"🏷️ <b>Tipo movimiento:</b> EDITAR\n"
        f"📊 <b>Stock anterior:</b> {stock_anterior} unidades\n"
        f"📊 <b>Stock nuevo:</b> {d.get('stock')} unidades\n"
        f"💰 <b>Precio:</b> {precio} COP\n"
        f"📍 <b>Bodega:</b> {d.get('bodega') or 'No especificada'}\n"
        f"⏰ <b>Hora:</b> {hora_colombia()}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"{alerta}"
        )
    enviar_telegram(mensaje)

    mysql.connection.commit()
    cursor.close()
    return jsonify({'ok': True})


@main.route('/api/productos/<int:id>', methods=['DELETE'])
def eliminar_producto(id):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

   
    cursor.execute("SELECT nombre, stock FROM productos WHERE id=%s", (id,))
    producto = cursor.fetchone()
    nombre_producto = producto['nombre'] if producto else ''
    stock_actual = producto['stock'] if producto else 0

   
    cursor.execute("""
        INSERT INTO movimientos 
        (id_producto, nombre_producto, tipo_movimiento, usuario, detalle, stock_anterior, stock_nuevo)
        VALUES (%s, %s, 'ELIMINAR', %s, %s, %s, %s)
    """, (
        id,
        nombre_producto,
        session.get('usuario'),
        "Producto eliminado del inventario",
        stock_actual,
        0
    ))

    
    cursor.execute("DELETE FROM productos WHERE id=%s", (id,))
    alerta = nivel_stock(0)
    mensaje = (
        f"🔴 <b>PRODUCTO ELIMINADO</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 <b>Producto:</b> {nombre_producto}\n"
        f"👤 <b>Usuario:</b> {session.get('usuario')}\n"
        f"🏷️ <b>Tipo movimiento:</b> ELIMINAR\n"
        f"📊 <b>Stock eliminado:</b> {stock_actual} unidades\n"
        f"📍 <b>Bodega:</b> No disponible (eliminado)\n"
        f"⏰ <b>Hora:</b> {hora_colombia()}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🔴 PRODUCTO FUERA DEL SISTEMA"
        )
    enviar_telegram(mensaje)

    mysql.connection.commit()
    cursor.close()
    return jsonify({'ok': True})

@main.route('/api/movimientos', methods=['GET'])
def get_movimientos():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT * FROM movimientos 
        ORDER BY fecha DESC 
        LIMIT 100
    """)
    resultado = cursor.fetchall()
    cursor.close()
    
    for row in resultado:
        if row.get('fecha'):
            fecha_local = row['fecha'] - timedelta(hours=5)
            row['fecha'] = fecha_local.strftime('%Y-%m-%d %I:%M:%S %p')
            
    return jsonify(resultado)
    


@main.route('/api/sesion')
def sesion():
    return jsonify({
        'activo':   bool(session.get('usuario')),
        'es_admin': session.get('rol') == 'admin',
        'es_asesor': session.get('rol') == 'asesor',
        'rol':      session.get('rol', '')
    })


# ============================================
# CLIENTES
# ============================================

@main.route('/api/clientes', methods=['GET'])
def get_clientes():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT * FROM clientes ORDER BY nombre ASC")
    resultado = cursor.fetchall()
    cursor.close()
    return jsonify(resultado)

@main.route('/api/clientes', methods=['POST'])
def crear_cliente():
    d = request.get_json()
    cursor = mysql.connection.cursor()
    try:
        cursor.execute("""
            INSERT INTO clientes (nombre, email, telefono, empresa)
            VALUES (%s, %s, %s, %s)
        """, (
            d.get('nombre'),
            d.get('email'),
            d.get('telefono'),
            d.get('empresa')
        ))
        mysql.connection.commit()
        cursor.close()
        return jsonify({'ok': True}), 201
    except Exception as e:
        cursor.close()
        return jsonify({'ok': False, 'mensaje': str(e)}), 400

# ============================================
# VENTAS
# ============================================

def generar_numero_factura():
    anio = datetime.now().strftime('%Y')
    numero = random.randint(1000, 9999)
    return f"FAC-{anio}-{numero}"

@main.route('/api/ventas', methods=['GET'])
def get_ventas():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT v.*, c.nombre as nombre_cliente
        FROM ventas v
        JOIN clientes c ON v.id_cliente = c.id
        ORDER BY v.fecha DESC
    """)
    resultado = cursor.fetchall()
    cursor.close()
    
    for row in resultado:
        if row.get('fecha'):
            # Restamos 5 horas (Hora Colombia)
            fecha_local = row['fecha'] - timedelta(hours=5)
            # Formato: YYYY-MM-DD HH:MM AM/PM
            row['fecha'] = fecha_local.strftime('%Y-%m-%d %I:%M %p')
            
    return jsonify(resultado))

@main.route('/api/ventas', methods=['POST'])
def crear_venta():
    d = request.get_json()
    id_cliente = d.get('id_cliente')
    items      = d.get('items')  # lista de {id_producto, cantidad, precio_unitario}
    notas      = d.get('notas', '')

    if not id_cliente or not items:
        return jsonify({'ok': False, 'mensaje': 'Datos incompletos'}), 400

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        # Calcular total
        total = sum(i['cantidad'] * i['precio_unitario'] for i in items)

        # Generar número de factura único
        numero_factura = generar_numero_factura()

        # Insertar venta
        cursor.execute("""
            INSERT INTO ventas (numero_factura, id_cliente, id_usuario, total, notas)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            numero_factura,
            id_cliente,
            session.get('usuario'),
            total,
            notas
        ))
        id_venta = cursor.lastrowid

        # Insertar detalle y descontar stock
        for item in items:
            id_producto     = item['id_producto']
            cantidad        = item['cantidad']
            precio_unitario = item['precio_unitario']
            subtotal        = cantidad * precio_unitario
            nombre_producto = item.get('nombre_producto', '')

            # Verificar stock suficiente
            cursor.execute("SELECT stock, nombre FROM productos WHERE id=%s", (id_producto,))
            producto = cursor.fetchone()

            if not producto or producto['stock'] < cantidad:
                raise Exception(f"Stock insuficiente para {nombre_producto}")

            # Insertar detalle
            cursor.execute("""
                INSERT INTO detalle_ventas
                (id_venta, id_producto, nombre_producto, cantidad, precio_unitario, subtotal)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (id_venta, id_producto, nombre_producto, cantidad, precio_unitario, subtotal))

            # Descontar stock
            cursor.execute("""
                UPDATE productos SET stock = stock - %s WHERE id = %s
            """, (cantidad, id_producto))

            # Registrar en movimientos
            cursor.execute("""
                INSERT INTO movimientos
                (id_producto, nombre_producto, tipo_movimiento, usuario, detalle, stock_anterior, stock_nuevo)
                VALUES (%s, %s, 'ELIMINAR', %s, %s, %s, %s)
            """, (
                id_producto,
                nombre_producto,
                session.get('usuario'),
                f"Venta #{numero_factura}",
                producto['stock'],
                producto['stock'] - cantidad
            ))

        mysql.connection.commit()

        # Notificación Telegram
        total_fmt = f"${int(total):,}".replace(',', '.')
        mensaje = (
            f"🧾 <b>NUEVA VENTA REGISTRADA</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📋 <b>Factura:</b> {numero_factura}\n"
            f"👤 <b>Asesor:</b> {session.get('usuario')}\n"
            f"🛍️ <b>Productos:</b> {len(items)}\n"
            f"💰 <b>Total:</b> {total_fmt} COP\n"
            f"⏰ <b>Hora:</b> {hora_colombia()}\n"
            f"━━━━━━━━━━━━━━━━━━━━"
        )
        enviar_telegram(mensaje)

        cursor.close()
        return jsonify({'ok': True, 'numero_factura': numero_factura, 'total': total}), 201

    except Exception as e:
        mysql.connection.rollback()
        cursor.close()
        return jsonify({'ok': False, 'mensaje': str(e)}), 400

@main.route('/api/ventas/<int:id>', methods=['GET'])
def get_venta_detalle(id):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT v.*, c.nombre as nombre_cliente, c.email, c.empresa, c.telefono
        FROM ventas v
        JOIN clientes c ON v.id_cliente = c.id
        WHERE v.id = %s
    """, (id,))
    venta = cursor.fetchone()

    if not venta:
        cursor.close()
        return jsonify({'ok': False, 'mensaje': 'Venta no encontrada'}), 404

    cursor.execute("""
        SELECT * FROM detalle_ventas WHERE id_venta = %s
    """, (id,))
    detalle = cursor.fetchall()
    cursor.close()

    if venta.get('fecha'):
        # Restamos 5 horas al detalle
        fecha_local = venta['fecha'] - timedelta(hours=5)
        venta['fecha'] = fecha_local.strftime('%Y-%m-%d %I:%M %p')

    return jsonify({'venta': venta, 'detalle': detalle})

# =======================================================
#   RUTAS DEL HISTORIAL DE VENTAS (USANDO EL BLUEPRINT)
# =======================================================

# 1. RUTA PARA VER LA PÁGINA WEB
@main.route('/historial-ventas')
def vista_historial_ventas():
    return render_template('HISTORIAL_VENTAS.html')


@main.route('/clientes')
def clientes_vista():
    if session.get('rol') not in ['admin', 'asesor']:
        return redirect(url_for('main.acceder'))
    return render_template('CLIENTES.html')


@main.route('/api/clientes/<int:id>', methods=['PUT'])
def actualizar_cliente(id):
    d = request.get_json()
    cursor = mysql.connection.cursor()
    try:
        cursor.execute("""
            UPDATE clientes 
            SET nombre=%s, email=%s, telefono=%s, empresa=%s
            WHERE id=%s
        """, (
            d.get('nombre'),
            d.get('email'),
            d.get('telefono'),
            d.get('empresa'),
            id
        ))
        mysql.connection.commit()
        cursor.close()
        return jsonify({'ok': True, 'mensaje': 'Cliente actualizado correctamente'})
    except Exception as e:
        cursor.close()
        return jsonify({'ok': False, 'mensaje': str(e)}), 400


from werkzeug.security import generate_password_hash

# ==========================================
# RUTAS DE LA VISTA Y API DE USUARIOS (SOLO ADMIN)
# ==========================================

@main.route('/usuarios')
def usuarios_vista():
    # Seguridad de la vista: Solo el admin entra aquí
    if session.get('rol') != 'admin':
        return redirect(url_for('main.inicio'))
    return render_template('USUARIOS.html')

@main.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401
        
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    # No enviamos las contraseñas al frontend por seguridad
    cursor.execute("SELECT id, usuario, rol, estado FROM usuarios ORDER BY usuario ASC")
    usuarios = cursor.fetchall()
    cursor.close()
    return jsonify(usuarios)

@main.route('/api/usuarios', methods=['POST'])
def crear_usuario():
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401
        
    d = request.get_json()
    usuario = d.get('usuario')
    password_plana = d.get('password')
    rol = d.get('rol')

    if not usuario or not password_plana or not rol:
        return jsonify({'ok': False, 'mensaje': 'Datos incompletos'}), 400

    password_hash = generate_password_hash(password_plana)
    cursor = mysql.connection.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO usuarios (usuario, password, rol, estado)
            VALUES (%s, %s, %s, 'activo')
        """, (usuario, password_hash, rol))
        mysql.connection.commit()
        cursor.close()
        return jsonify({'ok': True, 'mensaje': 'Usuario creado correctamente'})
    except Exception as e:
        cursor.close()
        return jsonify({'ok': False, 'mensaje': 'Error (¿El usuario ya existe?)'}), 400

@main.route('/api/usuarios/<int:id>', methods=['PUT'])
def actualizar_usuario(id):
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401
        
    d = request.get_json()
    rol = d.get('rol')
    estado = d.get('estado')
    nueva_password = d.get('password') # Puede venir vacía si no la quieren cambiar

    cursor = mysql.connection.cursor()
    try:
        if nueva_password:
            # Si el admin escribió una nueva clave, la encriptamos y la actualizamos
            password_hash = generate_password_hash(nueva_password)
            cursor.execute("""
                UPDATE usuarios SET rol=%s, estado=%s, password=%s WHERE id=%s
            """, (rol, estado, password_hash, id))
        else:
            # Si dejaron la clave vacía, solo actualizamos rol y estado
            cursor.execute("""
                UPDATE usuarios SET rol=%s, estado=%s WHERE id=%s
            """, (rol, estado, id))
            
        mysql.connection.commit()
        cursor.close()
        return jsonify({'ok': True, 'mensaje': 'Usuario actualizado'})
    except Exception as e:
        cursor.close()
        return jsonify({'ok': False, 'mensaje': str(e)}), 400

@main.route('/api/exportar/inventario', methods=['GET'])
def exportar_inventario():
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401

    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        # Adapté la consulta a las columnas de tu interfaz (stock en lugar de cantidad)
        cursor.execute("SELECT id, codigo, nombre, stock, costo, bodega FROM productos ORDER BY nombre ASC")
        productos = cursor.fetchall()
        cursor.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Inventario"

        # Encabezados de las columnas
        ws.append(['ID', 'Código', 'Producto', 'Stock Actual', 'Costo', 'Ubicación'])
        for celda in ws[1]:
            celda.font = openpyxl.styles.Font(bold=True)

        # Llenar datos
        for p in productos:
            ws.append([
                p.get('id'), p.get('codigo'), p.get('nombre'), 
                p.get('stock'), p.get('costo'), p.get('bodega')
            ])

        memoria = io.BytesIO()
        wb.save(memoria)
        memoria.seek(0)

        return send_file(
            memoria,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Inventario_PrintPack.xlsx'
        )

    except Exception as e:
        return jsonify({'ok': False, 'mensaje': str(e)}), 500


# ==========================================
# RUTA PARA EXPORTAR HISTORIAL DE VENTAS A EXCEL
# ==========================================
@main.route('/api/exportar/ventas', methods=['GET'])
def exportar_ventas():
    # Seguridad: Solo el admin puede descargar el reporte financiero
    if session.get('rol') != 'admin':
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401

    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        # Corregido: ON v.id_cliente = c.id
        # Mejora: Traemos v.numero_factura en lugar de v.id
        cursor.execute("""
            SELECT v.numero_factura, c.nombre AS nombre_cliente, v.fecha, v.total 
            FROM ventas v
            LEFT JOIN clientes c ON v.id_cliente = c.id
            ORDER BY v.fecha DESC
        """)
        ventas = cursor.fetchall()
        cursor.close()

        # Crear el archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        # Encabezados de las columnas
        encabezados = ['N° Factura', 'Cliente', 'Fecha de Venta', 'Total Pagado']
        ws.append(encabezados)

        # Formato de negrita para los encabezados
        for celda in ws[1]:
            celda.font = openpyxl.styles.Font(bold=True)

        # Llenar los datos fila por fila
        # Llenar los datos fila por fila
        for v in ventas:
            if v.get('fecha'):
                fecha_local = v.get('fecha') - timedelta(hours=5)
                fecha_limpia = fecha_local.strftime('%Y-%m-%d %I:%M %p')
            else:
                fecha_limpia = 'N/A'
            
            ws.append([
                v.get('numero_factura'), 
                v.get('nombre_cliente') or 'Cliente Mostrador', 
                fecha_limpia, 
                v.get('total')
            ])

        # Guardar en memoria temporal
        memoria = io.BytesIO()
        wb.save(memoria)
        memoria.seek(0)

        # Enviar el archivo para descarga
        return send_file(
            memoria,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Reporte_Ventas_PrintPack.xlsx'
        )

    except Exception as e:
        return jsonify({'ok': False, 'mensaje': str(e)}), 500


# ==========================================
# RUTA DEL PANEL PRINCIPAL (DASHBOARD)
# ==========================================
@main.route('/dashboard')
def dashboard():
    # Solo usuarios logueados pueden ver el panel
    if not session.get('rol'):
        return redirect(url_for('main.acceder'))
    return render_template('DASHBOARD.html')
